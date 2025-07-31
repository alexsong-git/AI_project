import requests
import boto3
from openpyxl import load_workbook
import json
import argparse
from typing import Dict, Any, List
from botocore.config import Config

# 创建 S3 客户端
s3 = boto3.client('s3', config=Config(signature_version='s3v4'))
bucket = 'ecms-user-email-message-dev'
def process_email_parsing(excel_file_path):
    """处理邮件解析流程的主函数"""
    url = 'http://order-email-parser:8080/parse-email'
    wb = load_workbook(excel_file_path)
    sheet = wb.active

    # 定义列索引 (1-based)
    html_path_col_index = 1
    subject_col_index = 2
    sender_col_index = 3
    html_url_col_index = 4
    request_col_index = 5  # 请求列
    response_col_index = 6  # 响应列 (第五列)
    expect_col_index = 7  # 预期结果列 (第六列)
    diff_col_index = 8  # 差异列 (第七列)


    # 处理所有请求
    process_all_requests(sheet, s3, url, html_path_col_index, html_url_col_index, subject_col_index,
                         sender_col_index, request_col_index, response_col_index)

    # 比较第五列和第六列的差异并写入第七列
    compare_columns_and_write_diffs(sheet, response_col_index, expect_col_index, diff_col_index)

    # 保存修改后的 Excel 文件
    wb.save(excel_file_path)
    print(f"处理完成，已保存结果到 {excel_file_path}")


def process_all_requests(sheet, s3_client, api_url, html_col, html_url_col, subject_col, sender_col,
                         request_col, response_col):
    """处理所有行的请求，从S3获取内容并发送到API"""
    for row_num in range(2, sheet.max_row + 1):
        try:
            html_path = sheet.cell(row=row_num, column=html_col).value
            subject = sheet.cell(row=row_num, column=subject_col).value
            sender = sheet.cell(row=row_num, column=sender_col).value

            request_body = build_request_body(s3_client, html_path, subject, sender)
            sheet.cell(row=row_num, column=request_col, value=json.dumps(request_body, indent=2))

            result = send_request(api_url, request_body)

            if isinstance(result, dict):
                sheet.cell(row=row_num, column=response_col, value=json.dumps(result, indent=2))
            else:
                sheet.cell(row=row_num, column=response_col, value=str(result))

        except Exception as e:
            error_msg = f'处理行 {row_num} 时发生异常: {str(e)}'
            sheet.cell(row=row_num, column=response_col, value=error_msg)
            print(error_msg)

        try:
            url = s3.generate_presigned_url(
                'get_object',
                Params={
                    'Bucket': bucket,
                    'Key': html_path,
                    'ResponseContentType': 'text/html'  # 强制浏览器渲染为HTML
                },
                ExpiresIn=3600
            )
            print(url)  # 输出可直接打开的临时链接
            sheet.cell(row=row_num, column=html_url_col, value=url)
        except Exception as e:
            sheet.cell(row=row_num, column=html_url_col, value='生成html链接失败')
            print('生成html链接失败')



def build_request_body(s3_client, html_path, subject, sender):
    """构建API请求体"""
    request_body = {
        "subject": subject,
        "sender": sender,
        "content": ""
    }

    if html_path:
        response_s3 = s3_client.get_object(Bucket=bucket, Key=html_path)
        request_body["content"] = response_s3['Body'].read().decode('utf-8')

    return request_body


def send_request(api_url, request_body):
    """发送请求到API并返回结果"""
    try:
        r = requests.post(api_url, json=request_body)
        if r.status_code == 200:
            return r.json()
        else:
            return f'请求失败，状态码: {r.status_code}'
    except Exception as e:
        return f'请求发生异常: {str(e)}'


def compare_columns_and_write_diffs(sheet, response_col, expect_col, diff_col):
    """对比响应列(第五列)和预期列(第六列)，将差异写入第七列"""
    for row_num in range(2, sheet.max_row + 1):
        try:
            # 获取第五列和第六列的内容
            response_str = str(sheet.cell(row=row_num, column=response_col).value or "")
            expect_str = str(sheet.cell(row=row_num, column=expect_col).value or "")

            # 解析JSON数据
            response_json = parse_json_safely(response_str)
            expect_json = parse_json_safely(expect_str)

            # 查找差异
            differences = find_json_differences(response_json, expect_json)

            # 写入差异结果
            if not differences:
                sheet.cell(row=row_num, column=diff_col, value="无差异")
            else:
                # 将差异列表转换为字符串
                diff_str = "; ".join(differences)
                # 处理过长的差异字符串，避免Excel单元格限制
                if len(diff_str) > 32767:  # Excel单元格字符限制
                    diff_str = diff_str[:32764] + "..."
                sheet.cell(row=row_num, column=diff_col, value=diff_str)

        except Exception as e:
            sheet.cell(row=row_num, column=diff_col, value=f"比较错误: {str(e)}")


def parse_json_safely(json_str: str) -> Dict[str, Any]:
    """安全地解析JSON字符串，处理可能的解析错误"""
    try:
        # 移除可能导致解析错误的多余空白字符
        json_str = json_str.strip()
        if not json_str:  # 处理空字符串
            return {}
        return json.loads(json_str)
    except json.JSONDecodeError:
        return {"解析错误": "无效的JSON格式"}
    except Exception:
        return {"解析错误": "处理JSON时出错"}


def find_json_differences(json1: Dict[str, Any], json2: Dict[str, Any]) -> List[str]:
    """找出两个JSON对象之间不同的字段名"""
    differences = []

    # 检查所有在第一个JSON中存在的键
    for key in json1:
        if key not in json2:
            differences.append(f"字段 '{key}' 仅存在于响应中")
        else:
            # 如果值都是字典，递归检查
            if isinstance(json1[key], dict) and isinstance(json2[key], dict):
                sub_diff = find_json_differences(json1[key], json2[key])
                # 为子字段添加父字段前缀
                differences.extend([f"{key}.{sub}" for sub in sub_diff])
            elif json1[key] != json2[key]:
                differences.append(f"字段 '{key}' 值不同")

    # 检查所有在第二个JSON中存在但第一个中不存在的键
    for key in json2:
        if key not in json1:
            differences.append(f"字段 '{key}' 仅存在于预期结果中")

    return differences


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description='邮件解析自动化工具，比较第五列和第六列差异')
    parser.add_argument('excel_path', help='Excel文件的完整路径')
    args = parser.parse_args()
    process_email_parsing(args.excel_path)
