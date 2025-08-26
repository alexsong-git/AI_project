import requests
import boto3
from openpyxl import load_workbook
import json
import sys
import os
from botocore.config import Config
from concurrent.futures import ThreadPoolExecutor, as_completed

# 创建 S3 客户端
s3 = boto3.client('s3', config=Config(signature_version='s3v4'))
bucket = 'ecms-user-email-message-dev'
# HTML文件存储根目录
HTML_ROOT_DIR = '/Users/alex/AI邮件解析'
# 线程池最大工作线程数
MAX_WORKERS = 5  # 可根据实际情况调整


def process_email_requests(excel_file_path, sheet_name):
    """处理邮件请求发送流程的主函数"""
    # 根据sheet名称创建对应的HTML目录
    html_dir = os.path.join(HTML_ROOT_DIR, f'html_body_{sheet_name.strip().lower()}')
    os.makedirs(html_dir, exist_ok=True)

    url = 'https://internal-api-dev.seel.com/order-email-parser/parse-email'
    wb = load_workbook(excel_file_path)

    # 创建存放JSON文件的目录（根据sheet名称区分）
    excel_dir = os.path.dirname(excel_file_path)
    json_dir = os.path.join(excel_dir, f'request_{sheet_name.strip().lower()}')
    # 创建存放响应JSON文件的目录
    response_dir = os.path.join(excel_dir, f'response_{sheet_name.strip().lower()}')
    os.makedirs(json_dir, exist_ok=True)
    os.makedirs(response_dir, exist_ok=True)

    # 根据指定的sheet名称获取工作表
    if sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        print(f"正在处理工作表: {sheet_name}")
    else:
        print(f"错误: 工作表 '{sheet_name}' 不存在于文件中")
        print(f"可用的工作表: {', '.join(wb.sheetnames)}")
        return

    # 定义列索引 (1-based)
    html_path_col_index = 1
    subject_col_index = 2
    sender_col_index = 3
    html_url_col_index = 4  # 现在存储本地HTML文件超链接
    request_col_index = 5  # 请求列（存储JSON文件链接）
    response_col_index = 6  # 响应列

    # 直接使用工作表名称作为数据类型
    data_type = sheet_name.strip().lower()
    data_type = ''.join([c if c.isalnum() or c == '_' else '_' for c in data_type])
    if len(data_type) > 15:
        data_type = data_type[:15]

    # 收集所有需要处理的行数据
    rows_to_process = []
    for row_num in range(2, sheet.max_row + 1):
        rows_to_process.append((
            row_num, sheet, s3, url, html_path_col_index, html_url_col_index,
            subject_col_index, sender_col_index, request_col_index,
            response_col_index, json_dir, response_dir, data_type, html_dir
        ))

    # 使用线程池处理所有行
    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
        # 提交所有任务
        futures = {executor.submit(process_single_row, *row_data): row_data for row_data in rows_to_process}

        # 等待所有任务完成
        for future in as_completed(futures):
            row_num = futures[future][0]
            try:
                future.result()
            except Exception as e:
                error_msg = f'线程处理行 {row_num} 时发生异常: {str(e)}'
                sheet.cell(row=row_num, column=response_col_index, value=error_msg)
                print(error_msg)

    # 保存修改后的 Excel 文件
    wb.save(excel_file_path)
    print(f"请求处理完成，已保存结果到 {excel_file_path}")
    print(f"请求JSON文件已保存到: {json_dir}")
    print(f"响应JSON文件已保存到: {response_dir}")
    print(f"HTML文件已下载到: {html_dir}")


def process_single_row(row_num, sheet, s3_client, api_url, html_col, html_url_col, subject_col, sender_col,
                       request_col, response_col, json_dir, response_dir, data_type, html_dir):
    """处理单行请求（供多线程调用）"""
    try:
        html_path = sheet.cell(row=row_num, column=html_col).value
        subject = sheet.cell(row=row_num, column=subject_col).value
        sender = sheet.cell(row=row_num, column=sender_col).value

        request_body = build_request_body(s3_client, html_path, subject, sender)

        # 使用工作表名作为类型生成文件名
        json_filename = f"request_{data_type}_row{row_num}.json"
        json_filepath = os.path.join(json_dir, json_filename)
        # 响应文件路径
        response_filename = f"response_{data_type}_row{row_num}.json"
        response_filepath = os.path.join(response_dir, response_filename)

        # 校验：如果subject为空则不写入request，清空原有内容
        if subject and str(subject).strip():
            # 保存请求体到JSON文件
            with open(json_filepath, 'w', encoding='utf-8') as f:
                json.dump(request_body, f, ensure_ascii=False, indent=2)

            # 在单元格插入JSON文件超链接
            sheet.cell(row=row_num, column=request_col).hyperlink = json_filepath
            sheet.cell(row=row_num, column=request_col).value = f"请求体 ({row_num}_{data_type})"
            sheet.cell(row=row_num, column=request_col).style = "Hyperlink"
        else:
            # 清空请求列链接和对应JSON文件
            sheet.cell(row=row_num, column=request_col).value = ""
            sheet.cell(row=row_num, column=request_col).hyperlink = None
            if os.path.exists(json_filepath):
                os.remove(json_filepath)

            # 清空响应相关内容
            sheet.cell(row=row_num, column=response_col).value = ""
            if os.path.exists(response_filepath):
                os.remove(response_filepath)
            return

        # 判断content是否为空，决定是否发起请求
        content = request_body.get("content")
        if content is not None and str(content).strip():
            result = send_request(api_url, request_body)

            # 保存响应结果到JSON文件
            with open(response_filepath, 'w', encoding='utf-8') as f:
                json.dump(result, f, ensure_ascii=False, indent=2)



            if isinstance(result, dict):
                sheet.cell(row=row_num, column=response_col, value=json.dumps(result, indent=2))
            else:
                sheet.cell(row=row_num, column=response_col, value=str(result))
        else:
            sheet.cell(row=row_num, column=response_col, value="内容为空，未发送请求")
            # 清空响应文件
            if os.path.exists(response_filepath):
                os.remove(response_filepath)


    except Exception as e:
        error_msg = f'处理行 {row_num} 时发生异常: {str(e)}'
        sheet.cell(row=row_num, column=response_col, value=error_msg)
        print(error_msg)

    try:
        # 如果html_path为空，不生成html_url
        if html_path:
            # 下载HTML文件到对应sheet的HTML目录
            local_html_path = download_html_from_s3(s3_client, html_path, row_num, data_type, html_dir)
            if local_html_path:
                # 设置本地HTML文件的超链接
                sheet.cell(row=row_num, column=html_url_col).hyperlink = local_html_path
                sheet.cell(row=row_num, column=html_url_col).value = f"本地HTML ({row_num}_{data_type})"
                sheet.cell(row=row_num, column=html_url_col).style = "Hyperlink"
            else:
                sheet.cell(row=row_num, column=html_url_col, value='HTML下载失败')
        else:
            # 清空原有可能存在的url
            sheet.cell(row=row_num, column=html_url_col, value="")
    except Exception as e:
        sheet.cell(row=row_num, column=html_url_col, value='处理HTML链接失败')
        print(f'处理行 {row_num} HTML链接失败: {str(e)}')


def download_html_from_s3(s3_client, s3_key, row_num, data_type, html_dir):
    """从S3下载HTML文件到对应sheet的HTML目录并返回本地路径"""
    try:
        # 生成本地文件名，避免重复
        filename = f"htmlbody_{data_type}_row{row_num}.html"
        local_path = os.path.join(html_dir, filename)

        # 下载S3文件到本地
        s3_client.download_file(bucket, s3_key, local_path)
        print(f"已下载HTML文件到: {local_path}")
        return local_path
    except Exception as e:
        print(f"下载HTML文件失败 (S3 key: {s3_key}): {str(e)}")
        return None


def build_request_body(s3_client, html_path, subject, sender):
    """构建API请求体"""
    request_body = {
        "subject": subject,
        "sender": sender,
        "content": ""
    }

    if html_path:
        response_s3 = s3_client.get_object(Bucket=bucket, Key=html_path)
        request_body["content"] = response_s3['Body'].read().decode('utf-8')

    return request_body


def send_request(api_url, request_body):
    """发送请求到API并返回结果"""
    try:
        r = requests.post(api_url, json=request_body)
        if r.status_code == 200:
            return r.json()
        else:
            return f'请求失败，状态码: {r.status_code}'
    except Exception as e:
        return f'请求发生异常: {str(e)}'


if __name__ == "__main__":
    if len(sys.argv) != 3:
        print("使用方法: python 请求接口.py <Excel文件路径> <工作表名称>")
        print("示例: python 请求接口.py /Users/alex/AI邮件解析/AI邮件解析自动化.xlsx ship")
        sys.exit(1)

    excel_path = sys.argv[1]
    sheet_name = sys.argv[2]
    process_email_requests(excel_path, sheet_name)