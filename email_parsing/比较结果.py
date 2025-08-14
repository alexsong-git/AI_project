from openpyxl import load_workbook
from openpyxl.styles import Alignment  # 导入Alignment类
import json
import argparse
from typing import Dict, Any, List


def process_json_comparison(excel_file_path, sheet_name):
    """处理JSON差异比较的主函数"""
    wb = load_workbook(excel_file_path)
    # 根据传入的sheet_name获取工作表
    sheet = wb[sheet_name]

    # 定义列索引 (1-based)
    response_col_index = 6  # 响应列
    expect_col_index = 7  # 预期结果列
    diff_col_index = 8  # 差异列

    # 比较响应列和预期列的差异并写入差异列
    compare_columns_and_write_diffs(sheet, response_col_index, expect_col_index, diff_col_index)

    # 保存修改后的 Excel 文件
    wb.save(excel_file_path)
    print(f"差异分析完成，已保存结果到 {excel_file_path}")


def compare_columns_and_write_diffs(sheet, response_col, expect_col, diff_col):
    """对比响应列和预期列，将差异写入差异列"""
    for row_num in range(2, sheet.max_row + 1):
        try:
            # 获取响应列和预期列的内容
            response_str = str(sheet.cell(row=row_num, column=response_col).value or "")
            expect_str = str(sheet.cell(row=row_num, column=expect_col).value or "")

            # 如果预期结果为空，不做对比
            if not expect_str.strip():
                sheet.cell(row=row_num, column=diff_col, value="预期结果为空，不进行对比")
                continue

            # 解析JSON数据
            response_json = parse_json_safely(response_str)
            expect_json = parse_json_safely(expect_str)

            # 查找差异（使用优化后的列表比较方法）
            differences = find_json_differences(response_json, expect_json)

            # 写入差异结果
            if not differences:
                sheet.cell(row=row_num, column=diff_col, value="无差异")
            else:
                # 将差异列表用换行符连接，实现Excel换行
                diff_str = "\n".join(differences)
                # 处理过长的差异字符串，避免Excel单元格限制
                if len(diff_str) > 32767:  # Excel单元格字符限制
                    diff_str = diff_str[:32764] + "..."
                # 设置单元格并开启自动换行（修复弃用警告）
                cell = sheet.cell(row=row_num, column=diff_col, value=diff_str)
                cell.alignment = Alignment(wrapText=True)  # 直接创建新的Alignment对象

        except Exception as e:
            sheet.cell(row=row_num, column=diff_col, value=f"比较错误: {str(e)}")


def parse_json_safely(json_str: str) -> Dict[str, Any]:
    """安全地解析JSON字符串，处理可能的解析错误"""
    try:
        # 移除可能导致解析错误的多余空白字符
        json_str = json_str.strip()
        if not json_str:  # 处理空字符串
            return {}
        return json.loads(json_str)
    except json.JSONDecodeError:
        return {"解析错误": "无效的JSON格式"}
    except Exception:
        return {"解析错误": "处理JSON时出错"}


def find_json_differences(json1: Any, json2: Any, parent_key: str = "") -> List[str]:
    """找出两个JSON对象（支持字典和列表）之间不同的字段名，精确到列表索引和子字段"""
    differences = []
    current_key = parent_key if parent_key else ""

    # 处理列表类型
    if isinstance(json1, list) and isinstance(json2, list):
        # 列表长度不同的情况
        if len(json1) != len(json2):
            differences.append(f"{current_key} 列表长度不同: 响应有{len(json1)}个元素，预期有{len(json2)}个元素")
            return differences

        # 比较列表中每个元素
        for i in range(len(json1)):
            # 为列表元素构建键名（如 line_items[0]）
            list_item_key = f"{current_key}[{i}]" if current_key else f"[{i}]"
            # 递归比较列表元素
            item_diffs = find_json_differences(json1[i], json2[i], list_item_key)
            differences.extend(item_diffs)
        return differences

    # 处理非列表但类型不同的情况
    if type(json1) != type(json2):
        differences.append(f"{current_key} 类型不同: 响应是{type(json1).__name__}，预期是{type(json2).__name__}")
        return differences

    # 处理字典类型（递归比较子字段）
    if isinstance(json1, dict) and isinstance(json2, dict):
        # 检查所有在第一个JSON中存在的键
        for key in json1:
            new_key = f"{current_key}.{key}" if current_key else key
            if key not in json2:
                differences.append(f"字段 '{new_key}' 仅存在于响应中")
            else:
                # 递归比较子字段
                sub_diffs = find_json_differences(json1[key], json2[key], new_key)
                differences.extend(sub_diffs)

        # 检查所有在第二个JSON中存在但第一个中不存在的键
        for key in json2:
            if key not in json1:
                new_key = f"{current_key}.{key}" if current_key else key
                differences.append(f"字段 '{new_key}' 仅存在于预期结果中")
        return differences

    # 处理基本类型（字符串、数字、布尔等）
    if json1 != json2:
        differences.append(f"{current_key} 值不同: 响应为{json1!r}，预期为{json2!r}")

    return differences


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description='JSON响应与预期结果差异分析工具')
    parser.add_argument('excel_path', help='Excel文件的完整路径')
    parser.add_argument('sheet_name', help='需要处理的工作表名称')
    args = parser.parse_args()
    process_json_comparison(args.excel_path, args.sheet_name)