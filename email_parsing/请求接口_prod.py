import requests
import boto3
from openpyxl import load_workbook
import json
import sys
import os
from botocore.config import Config
from concurrent.futures import ThreadPoolExecutor, as_completed
import base64
import hashlib
from Crypto.Cipher import AES
from datetime import datetime  # 新增：导入datetime处理日期

# 创建 S3 客户端
s3 = boto3.client('s3', config=Config(signature_version='s3v4'))
#bucket = 'ecms-user-email-message-dev'
bucket = 'ecms-user-email-message'
# HTML文件存储根目录
HTML_ROOT_DIR = '/Users/alex/AI邮件解析'
# 主线程池最大工作线程数
MAX_WORKERS = 30  # 可根据实际情况调整
# HTML下载专用线程池大小
HTML_DOWNLOAD_WORKERS = 30  # 可根据网络情况调整

# 添加解密相关的常量和函数
SECRET_KEY = 'seel-fetch-email-secret'


def process_key(key: str) -> bytes:
    # 直接处理为256位密钥（32字节）
    return hashlib.sha256(key.encode('utf-8')).digest()


def decode_base64(base64_data: str) -> bytes:
    """
    Decode URL-safe base64 string to bytes, handling padding.
    """
    base64_str = base64_data.replace('-', '+').replace('_', '/')
    padding = len(base64_str) % 4
    if padding == 2:
        base64_str += "=="
    elif padding == 3:
        base64_str += "="
    return base64.b64decode(base64_str)


def symmetric_decrypt_with_base64_decode(key: str, encrypted_value: str) -> str:
    """
    解密并Base64解码
    :param key: 加密密钥
    :param encrypted_value: 十六进制的AES加密字符串
    :return: 解密后的明文字符串
    """
    try:
        key_bytes = process_key(key)
        cipher = AES.new(key_bytes, AES.MODE_ECB)
        # 先将hex字符串转为字节
        encrypted_bytes = bytes.fromhex(encrypted_value)
        decrypted = cipher.decrypt(encrypted_bytes)
        # 去除PKCS7填充
        pad_len = decrypted[-1]
        decrypted = decrypted[:-pad_len]
        # base64解码得到原始内容
        decoded = decode_base64(decrypted.decode('utf-8'))
        return decoded.decode('utf-8')
    except Exception as e:
        # 这里可以用logging模块替换
        print(f"Error during symmetric decryption: {e}")
        return None


def process_email_requests(excel_file_path, sheet_name):
    """处理邮件请求发送流程的主函数"""
    # 根据sheet名称创建对应的HTML目录
    html_dir = os.path.join(HTML_ROOT_DIR, f'html_body_{sheet_name.strip().lower()}')
    os.makedirs(html_dir, exist_ok=True)

    # url = 'https://internal-api-dev.seel.com/order-email-parser/parse-email'
    url = 'http://order-email-parser:8080/parse-email'
    wb = load_workbook(excel_file_path)

    # 创建存放JSON文件的目录（根据sheet名称区分）
    excel_dir = os.path.dirname(excel_file_path)
    json_dir = os.path.join(excel_dir, f'request_{sheet_name.strip().lower()}')
    # 创建存放响应JSON文件的目录
    response_dir = os.path.join(excel_dir, f'response_{sheet_name.strip().lower()}')
    os.makedirs(json_dir, exist_ok=True)
    os.makedirs(response_dir, exist_ok=True)

    # 根据指定的sheet名称获取工作表
    if sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        print(f"正在处理工作表: {sheet_name}")
    else:
        print(f"错误: 工作表 '{sheet_name}' 不存在于文件中")
        print(f"可用的工作表: {', '.join(wb.sheetnames)}")
        return

    # 定义列索引 (1-based)
    html_path_col_index = 1
    subject_col_index = 2
    sender_col_index = 3
    html_url_col_index = 4  # 现在存储本地HTML文件超链接
    request_col_index = 5  # 请求列（存储JSON文件链接）
    response_col_index = 6  # 响应列
    date_col_index = 9  # 收到邮件时间

    # 直接使用工作表名称作为数据类型
    data_type = sheet_name.strip().lower()
    data_type = ''.join([c if c.isalnum() or c == '_' else '_' for c in data_type])
    if len(data_type) > 15:
        data_type = data_type[:15]

    # 收集所有需要处理的行数据（跳过隐藏行）
    rows_to_process = []
    for row_num in range(2, sheet.max_row + 1):
        # 检查行是否隐藏（openpyxl中row_dimensions的hidden属性）
        if sheet.row_dimensions[row_num].hidden:
            continue
        rows_to_process.append((
            row_num, sheet, s3, url, html_path_col_index, html_url_col_index,
            subject_col_index, sender_col_index, request_col_index,
            response_col_index, json_dir, response_dir, data_type, html_dir,
            date_col_index  # 传递日期列索引
        ))

    # 创建HTML下载专用线程池
    html_executor = ThreadPoolExecutor(max_workers=HTML_DOWNLOAD_WORKERS)
    html_futures = []

    # 使用主线程池处理所有行
    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
        # 提交所有任务
        futures = {executor.submit(process_single_row, *row_data, html_executor, html_futures): row_data for row_data in
                   rows_to_process}

        # 等待所有任务完成
        for future in as_completed(futures):
            row_num = futures[future][0]
            try:
                future.result()
            except Exception as e:
                error_msg = f'线程处理行 {row_num} 时发生异常: {str(e)}'
                sheet.cell(row=row_num, column=response_col_index, value=error_msg)
                print(error_msg)

    # 等待所有HTML下载任务完成
    print("等待所有HTML下载任务完成...")
    for future in as_completed(html_futures):
        try:
            future.result()
        except Exception as e:
            print(f'HTML下载任务异常: {str(e)}')
    html_executor.shutdown()

    # 保存修改后的 Excel 文件
    wb.save(excel_file_path)
    print(f"请求处理完成，已保存结果到 {excel_file_path}")
    print(f"请求JSON文件已保存到: {json_dir}")
    print(f"响应JSON文件已保存到: {response_dir}")
    print(f"HTML文件已下载到: {html_dir}")


def process_single_row(row_num, sheet, s3_client, api_url, html_col, html_url_col, subject_col, sender_col,
                       request_col, response_col, json_dir, response_dir, data_type, html_dir, date_col_index,
                       html_executor, html_futures):
    """处理单行请求（供多线程调用）"""
    try:
        # 获取日期列数据并处理（核心修改：仅保留年月日）
        date_value = sheet.cell(row=row_num, column=date_col_index).value
        date_str = "no_date"  # 默认值

        if date_value is not None:
            # 处理datetime对象
            if isinstance(date_value, datetime):
                # 直接格式化成年月日（不含时分秒）
                date_str = date_value.strftime("%Y%m%d")
            else:
                # 处理字符串格式日期
                date_str_raw = str(date_value).strip()
                # 尝试从字符串中提取年月日（支持常见格式）
                try:
                    # 尝试解析包含时分秒的格式（如"2023-10-05 14:30:00"）
                    date_obj = datetime.strptime(date_str_raw, "%Y-%m-%d %H:%M:%S")
                    date_str = date_obj.strftime("%Y%m%d")
                except ValueError:
                    try:
                        # 尝试仅日期格式（如"2023-10-05"）
                        date_obj = datetime.strptime(date_str_raw, "%Y-%m-%d")
                        date_str = date_obj.strftime("%Y%m%d")
                    except ValueError:
                        # 提取所有数字后取前8位（确保只到日）
                        date_digits = ''.join(filter(str.isdigit, date_str_raw))
                        if len(date_digits) >= 8:
                            date_str = date_digits[:8]  # 取前8位（YYYYMMDD）
                        else:
                            date_str = "invalid_date"

        html_path = sheet.cell(row=row_num, column=html_col).value
        subject = sheet.cell(row=row_num, column=subject_col).value
        sender = sheet.cell(row=row_num, column=sender_col).value

        request_body = build_request_body(s3_client, html_path, subject, sender)

        # 使用工作表名作为类型生成文件名，添加日期后缀（仅年月日）
        json_filename = f"request_{data_type}_row{row_num}_{date_str}.json"
        json_filepath = os.path.join(json_dir, json_filename)
        # 响应文件路径，添加日期后缀（仅年月日）
        response_filename = f"response_{data_type}_row{row_num}_{date_str}.json"
        response_filepath = os.path.join(response_dir, response_filename)

        # 校验：如果subject为空则不写入request，清空原有内容
        if subject and str(subject).strip():
            # 保存请求体到JSON文件
            with open(json_filepath, 'w', encoding='utf-8') as f:
                json.dump(request_body, f, ensure_ascii=False, indent=2)

            # 在单元格插入JSON文件超链接
            sheet.cell(row=row_num, column=request_col).hyperlink = json_filepath
            sheet.cell(row=row_num, column=request_col).value = f"请求体 ({row_num}_{data_type})"
            sheet.cell(row=row_num, column=request_col).style = "Hyperlink"
        else:
            # 清空请求列链接和对应JSON文件
            sheet.cell(row=row_num, column=request_col).value = ""
            sheet.cell(row=row_num, column=request_col).hyperlink = None
            if os.path.exists(json_filepath):
                os.remove(json_filepath)

            # 清空响应相关内容
            sheet.cell(row=row_num, column=response_col).value = ""
            if os.path.exists(response_filepath):
                os.remove(response_filepath)
            return

        # 判断content是否为空，决定是否发起请求
        content = request_body.get("content")
        if content is not None and str(content).strip():
            result = send_request(api_url, request_body)

            # 保存响应结果到JSON文件
            with open(response_filepath, 'w', encoding='utf-8') as f:
                json.dump(result, f, ensure_ascii=False, indent=2)

            if isinstance(result, dict):
                sheet.cell(row=row_num, column=response_col, value=json.dumps(result, indent=2))
            else:
                sheet.cell(row=row_num, column=response_col, value=str(result))
        else:
            sheet.cell(row=row_num, column=response_col, value="内容为空，未发送请求")
            # 清空响应文件
            if os.path.exists(response_filepath):
                os.remove(response_filepath)

    except Exception as e:
        error_msg = f'处理行 {row_num} 时发生异常: {str(e)}'
        sheet.cell(row=row_num, column=response_col, value=error_msg)
        print(error_msg)

    try:
        # 如果html_path为空，不生成html_url
        if html_path:
            # 提交HTML下载任务到专用线程池，传递日期参数
            future = html_executor.submit(
                handle_html_download,
                s3_client, html_path, row_num, data_type, html_dir,
                sheet, html_url_col, date_str
            )
            html_futures.append(future)
        else:
            # 清空原有可能存在的url
            sheet.cell(row=row_num, column=html_url_col).value = ""
    except Exception as e:
        sheet.cell(row=row_num, column=html_url_col).value = '处理HTML链接失败'
        print(f'处理行 {row_num} HTML链接失败: {str(e)}')


def handle_html_download(s3_client, s3_key, row_num, data_type, html_dir, sheet, html_url_col, date_str):
    """处理HTML下载的专用函数，在独立线程池中执行"""
    try:
        local_html_path = download_html_from_s3(s3_client, s3_key, row_num, data_type, html_dir, date_str)
        if local_html_path:
            # 设置本地HTML文件的超链接
            sheet.cell(row=row_num, column=html_url_col).hyperlink = local_html_path
            sheet.cell(row=row_num, column=html_url_col).value = f"本地HTML ({row_num}_{data_type})"
            sheet.cell(row=row_num, column=html_url_col).style = "Hyperlink"
        else:
            sheet.cell(row=row_num, column=html_url_col).value = 'HTML下载失败'
    except Exception as e:
        sheet.cell(row=row_num, column=html_url_col).value = 'HTML下载异常'
        print(f'行 {row_num} HTML下载异常: {str(e)}')


def download_html_from_s3(s3_client, s3_key, row_num, data_type, html_dir, date_str):
    """从S3下载HTML文件到对应sheet的HTML目录并返回本地路径"""
    try:
        # 生成本地文件名，添加日期后缀（仅年月日）
        filename = f"htmlbody_{data_type}_row{row_num}_{date_str}.html"
        local_path = os.path.join(html_dir, filename)

        # 下载S3文件到本地
        s3_client.download_file(bucket, s3_key, local_path)

        # 读取下载的文件内容并尝试解密
        with open(local_path, 'r', encoding='utf-8') as f:
            encrypted_content = f.read()

        # 尝试解密内容
        decrypted_content = symmetric_decrypt_with_base64_decode(SECRET_KEY, encrypted_content)

        if decrypted_content:
            # 将解密后的内容写回文件
            with open(local_path, 'w', encoding='utf-8') as f:
                f.write(decrypted_content)
            print(f"已下载并解密HTML文件到: {local_path}")
        else:
            print(f"解密失败，保留原始加密内容: {local_path}")

        return local_path
    except Exception as e:
        print(f"下载或解密HTML文件失败 (S3 key: {s3_key}): {str(e)}")
        return None


def build_request_body(s3_client, html_path, subject, sender):
    """构建API请求体"""
    request_body = {
        "subject": subject,
        "sender": sender,
        "content": ""
    }

    if html_path:
        response_s3 = s3_client.get_object(Bucket=bucket, Key=html_path)
        encrypted_content = response_s3['Body'].read().decode('utf-8')

        # 尝试解密内容
        decrypted_content = symmetric_decrypt_with_base64_decode(SECRET_KEY, encrypted_content)

        if decrypted_content:
            request_body["content"] = decrypted_content
        else:
            request_body["content"] = encrypted_content  # 如果解密失败，使用原始内容

    return request_body


def send_request(api_url, request_body):
    """发送请求到API并返回结果"""
    try:
        r = requests.post(api_url, json=request_body)
        if r.status_code == 200:
            return r.json()
        else:
            return f'请求失败，状态码: {r.status_code}'
    except Exception as e:
        return f'请求发生异常: {str(e)}'


if __name__ == "__main__":
    if len(sys.argv) != 3:
        print("使用方法: python 请求接口.py <Excel文件路径> <工作表名称>")
        print("示例: python 请求接口.py /Users/alex/AI邮件解析/AI邮件解析自动化.xlsx ship")
        sys.exit(1)

    excel_path = sys.argv[1]
    sheet_name = sys.argv[2]
    process_email_requests(excel_path, sheet_name)