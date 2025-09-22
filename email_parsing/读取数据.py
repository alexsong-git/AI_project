import psycopg2
from openpyxl import load_workbook
import sys

# 数据库连接参数
db_params = {
    "host": "seel-dev.cluster-c4pawlij3cb6.us-east-2.rds.amazonaws.com",
    "database": "ecm_saas",
    "user": "postgres",
    "password": "LlBu8KlK1QKNyZaUqgPv",
    "port": "5432"
}

# 查询语句字典
queries = {
    "order": """
    SELECT "html_body", "subject", "sender" 
    FROM "EmailMessage" 
    WHERE "email" IN ('minji.kim10101@gmail.com','panxiuyan749@gmail.com','jie.liu@seel.com','zhe.jiang@seel.com','xingchen.alexsong@gmail.com','kai.li@seel.com','hao.zhou@seel.com')
    AND "business_type" = 'ORDER' 
    AND ("subject" ILIKE '%confir%' OR "subject" ILIKE '%receive%' OR "subject" ILIKE '%new%' OR "subject" ILIKE '%processe%') 
    AND "datetime" BETWEEN '2025-07-30 00:00:00+00' AND '2025-07-30 23:00:00+00'
    ORDER BY "email", "created_ts" DESC;
    """,
    "ship": """
    SELECT "html_body", "subject", "sender" 
    FROM "EmailMessage" 
    WHERE "email" IN ('minji.kim10101@gmail.com','panxiuyan749@gmail.com','jie.liu@seel.com','zhe.jiang@seel.com','xingchen.alexsong@gmail.com','kai.li@seel.com','hao.zhou@seel.com')
    AND "business_type" = 'ORDER' 
    AND ("subject" ILIKE '%ship%' OR "subject" ILIKE '%way%') 
    AND "datetime" BETWEEN '2025-07-30 00:00:00+00' AND '2025-08-05 23:00:00+00'
    ORDER BY "email", "created_ts" DESC;
    """,
    "cancel": """
    SELECT "html_body", "subject", "sender" 
    FROM "EmailMessage" 
    WHERE "email" IN ('minji.kim10101@gmail.com','panxiuyan749@gmail.com','jie.liu@seel.com','zhe.jiang@seel.com','xingchen.alexsong@gmail.com','kai.li@seel.com','hao.zhou@seel.com')
    AND "business_type" = 'ORDER' 
    AND ("subject" ILIKE '%cancel%') 
    AND "datetime" BETWEEN '2025-07-30 00:00:00+00' AND '2025-08-05 23:00:00+00'
    ORDER BY "email", "created_ts" DESC;
    """,
    "return": """
    SELECT "html_body", "subject", "sender"
    FROM "EmailMessage" WHERE "email" = 'songyuchen93@gmail.com' AND "business_type" = 'RETURN' order by "created_ts" DESC ;
    """,
    "test": """
    SELECT "html_body", "subject", "sender" 
    FROM "EmailMessage" 
    WHERE "email" IN ('calvinkxu@gmail.com','kswindall13@gmail.com','nfaggioli@gmail.com','spirz.chris@gmail.com')
    AND ("business_type" != 'RETURN' OR "business_type" IS NULL)
    AND "datetime" > '2025-08-27 00:00:00+00'
    ORDER BY "email", "created_ts" DESC;
    """,
    "all": """
    SELECT "html_body", "subject", "sender" 
    FROM "EmailMessage" 
    WHERE 1=1
    AND ("business_type" != 'RETURN' OR "business_type" IS NULL)
    AND "datetime" > '2025-08-20 00:00:00+00'
    ORDER BY "email", "created_ts" DESC;
    """
}


def get_next_visible_row(sheet, start_row):
    """获取从起始行开始的下一个可见行"""
    current_row = start_row
    while True:
        # 如果行号超过最大行，视为可见（新建行）
        if current_row > sheet.max_row:
            return current_row
        # 检查行是否隐藏（默认未隐藏）
        if not sheet.row_dimensions.get(current_row, None) or not sheet.row_dimensions[current_row].hidden:
            return current_row
        current_row += 1


def export_email_data_to_excel(excel_path, query_type):
    """
    将数据库中的邮件数据导出到指定路径的Excel文件的对应工作表
    只操作前3列，不影响其他列数据
    """
    if query_type not in queries:
        print(f"错误：无效的查询类型 '{query_type}'，必须是 'order'、'ship' 或 'cancel'")
        return

    connection = None
    try:
        connection = psycopg2.connect(** db_params)
        cursor = connection.cursor()

        cursor.execute(queries[query_type])
        results = cursor.fetchall()

        workbook = load_workbook(excel_path)

        if query_type not in workbook.sheetnames:
            workbook.create_sheet(title=query_type)

        sheet = workbook[query_type]

        # 只清空前3列的可见行数据（保留表头），不影响其他列
        max_row = sheet.max_row
        for row in range(2, max_row + 1):
            # 只清理可见行
            if not sheet.row_dimensions.get(row, None) or not sheet.row_dimensions[row].hidden:
                for col in range(1, 4):  # 只操作第1-3列
                    sheet.cell(row=row, column=col).value = None

        # 写入新数据（只写入前3列，跳过隐藏行）
        current_row = 2
        for row_data in results:
            # 获取下一个可见行
            write_row = get_next_visible_row(sheet, current_row)
            # 只写入查询结果涉及的3列数据
            for col_index in range(3):
                sheet.cell(row=write_row, column=col_index + 1, value=row_data[col_index])
            current_row = write_row + 1

        workbook.save(excel_path)
        print(f"{query_type} 数据已成功写入 Excel 文件的 {query_type} 工作表，不影响其他列数据。")

    except (Exception, psycopg2.Error) as error:
        print("连接数据库或执行查询时出现错误:", error)
    finally:
        if connection:
            cursor.close()
            connection.close()
            print("数据库连接已关闭。")


# 从命令行接收参数
if __name__ == "__main__":
    if len(sys.argv) != 3:
        print("使用方法: python test.py <excel文件路径> <查询类型(order/ship/cancel)>")
        sys.exit(1)

    excel_file_path = sys.argv[1]
    query_type = sys.argv[2]

    export_email_data_to_excel(excel_file_path, query_type)